import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# 🔹 Configuración de conexión a SQL Server
server = 'PCMANU\\SQLEXPRESS'  # O 'localhost\SQLEXPRESS'
database = 'Bar'
username = 'user1'
password = '123456789'

conn_string = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'

# 🔹 Conectar a SQL Server y obtener datos
def fetch_data():
    try:
        conn = pyodbc.connect(conn_string)
        cursor = conn.cursor()

        query = """
        SELECT Beer.BeerID, Beer.Name AS BeerName, Brand.Name AS BrandName 
        FROM Beer 
        INNER JOIN Brand ON Beer.BrandID = Brand.BrandID
        """
        cursor.execute(query)
        
        # Obtener nombres de columnas
        columns = [column[0] for column in cursor.description]

         # Convertir filas a listas para evitar errores de forma
        data = [list(row) for row in cursor.fetchall()]
        
        cursor.close()
        conn.close()

        # Verificar si hay datos antes de crear el DataFrame
        if not data:
            print("No se encontraron resultados en la consulta.")
            return pd.DataFrame(columns=columns)

        return pd.DataFrame(data, columns=columns)

    except Exception as e:
        print("Error en la conexión:", e)
        return pd.DataFrame()

# 🔹 Exportar datos a Excel con formato
def export_to_excel(df):
    if df.empty:
        print("No hay datos para exportar.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Beers"

    # Agregar encabezados con formato
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(left=Side(style='thin', color="000000"), 
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000"))

    ws.append(df.columns.tolist())  # Agregar encabezados

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Agregar datos con bordes
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)
        for c_idx, cell in enumerate(ws[r_idx], start=1):
            cell.border = border

    # Guardar archivo Excel
    wb.save("BeersList.xlsx")
    print("Archivo 'BeersList.xlsx' generado correctamente.")

# 🔹 Ejecutar el proceso
if __name__ == "__main__":
    data = fetch_data()
    export_to_excel(data)
